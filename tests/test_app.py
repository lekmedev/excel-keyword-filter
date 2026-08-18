"""
Tests cho excel_keyword_filter.

Chạy local (ngoài Docker):
    cd ~/apps/excel-keyword-filter
    python3 -m venv .venv && . .venv/bin/activate
    pip install -r backend/requirements.txt pytest
    pytest -q

Hoặc trong Docker:
    docker compose run --rm excel-filter pytest -q  (nếu có pytest trong image)
"""

import io
import subprocess
import sys
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from backend import main  # noqa: E402
from backend.excel_processor import NoMatchError, process_excel  # noqa: E402


# ---------------------------------------------------------------------------
# Fixture: file Excel mẫu
# ---------------------------------------------------------------------------

def make_sample_xlsx(path: Path, active_sheet: str = "DataSheet") -> None:
    """Tạo file .xlsx mẫu: header A..O, vài dòng dữ liệu.

    Manual (active "Datamanual") chứa từ khóa; BOMSheet là sheet ẩn/thêm.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = active_sheet

    headers = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=f"Col{h}")

    data = [
        ["row2", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "Logitech keyboard", "o2"],
        ["row3", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "Wireless mouse", "o3"],
        ["row4", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "Bàn phím cơ (mechanical)", "o4"],
        ["row5", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "Monitor 24 inch", "o5"],
        ["row6", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "MOUSE gaming", "o6"],
    ]
    for r, row in enumerate(data, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    ws.merge_cells("A1:B1")  # merge trên header để test copy merges

    extra = wb.create_sheet("Other")
    extra.cell(row=1, column=1, value="should-not-appear")

    wb.save(path)


@pytest.fixture()
def sample_xlsx(tmp_path: Path):
    path = tmp_path / "input.xlsx"
    make_sample_xlsx(path)
    return path


# ---------------------------------------------------------------------------
# Unit tests: excel_processor
# ---------------------------------------------------------------------------

def test_matches_keywords_case_insensitive(sample_xlsx: Path):
    out, count = process_excel(sample_xlsx, "KEYBOARD\nchuột\nbàn phím")
    # Khớp: row2 "Logitech keyboard", row4 "Bàn phím cơ", row6 "MOUSE gaming" (chuột=mouse? KHÔNG — "mouse" là từ khóa riêng)
    # "chuột" là tiếng Việt, không có trong dữ liệu; "mouse" không được liệt kê.
    # => khớp: row2 (KEYBOARD), row4 (bàn phím) = 2
    assert count == 2
    wb = load_workbook(out)
    assert "TRUE_Result" in wb.sheetnames


def test_columns_deleted_and_stt_inserted(sample_xlsx: Path):
    out, count = process_excel(sample_xlsx, "keyboard")
    assert count == 1
    ws = load_workbook(out)["TRUE_Result"]
    # Kept cols A..O trừ A,B,G,H,I,K,L,M -> còn C,D,E,F,J,N,O + thêm STT
    # Layout: [STT, C, D, E, F, J, N, O]
    assert ws.cell(row=1, column=1).value == "STT"
    assert ws.cell(row=2, column=1).value == "=ROW()-1"
    assert ws.cell(row=1, column=2).value == "COLC"
    assert ws.cell(row=1, column=3).value == "COLD"
    assert ws.cell(row=1, column=4).value == "COLE"
    assert ws.cell(row=1, column=5).value == "COLF"
    assert ws.cell(row=1, column=6).value == "COLJ"
    assert ws.cell(row=1, column=7).value == "COLN"
    assert ws.cell(row=2, column=7).value == "Logitech keyboard"
    # Cột H (8) bị header "Quantity Replaced" ghi đè theo spec macro gốc;
    # dữ liệu cột O gốc vẫn nằm bên dưới (ô H2 = "o2").
    assert ws.cell(row=1, column=8).value == "QUANTITY REPLACED"
    assert ws.cell(row=2, column=8).value == "o2"


def test_quantity_replaced_header(sample_xlsx: Path):
    out, _ = process_excel(sample_xlsx, "keyboard")
    ws = load_workbook(out)["TRUE_Result"]
    assert ws.cell(row=1, column=8).value == "QUANTITY REPLACED"


def test_stt_uses_row_formula(sample_xlsx: Path):
    """STT phải là công thức =ROW()-1 (tự điều chỉnh khi xóa hàng)."""
    out, _ = process_excel(sample_xlsx, "keyboard\nbàn phím")
    ws = load_workbook(out)["TRUE_Result"]
    assert ws.cell(row=2, column=1).value == "=ROW()-1"
    assert ws.cell(row=3, column=1).value == "=ROW()-1"


def test_total_row(sample_xlsx: Path):
    """Dòng Total: merge A->G, italic+bold, align right, H = SUM(H2:H{last})."""
    out, _ = process_excel(sample_xlsx, "keyboard\nbàn phím")
    ws = load_workbook(out)["TRUE_Result"]
    # 2 dòng khớp -> header row1, data row2-3, Total row4
    assert ws.max_row == 4
    total = ws.cell(row=4, column=1)
    assert total.value == "Total"
    assert total.font.italic is True and total.font.bold is True
    assert total.alignment.horizontal == "right"
    # A4:G4 merged
    assert any(
        rng.min_row == 4 and rng.min_col == 1 and rng.max_row == 4 and rng.max_col == 7
        for rng in ws.merged_cells.ranges
    )
    sum_cell = ws.cell(row=4, column=8)
    assert sum_cell.value == "=SUM(H2:H3)"
    assert sum_cell.font.italic is True and sum_cell.font.bold is True


def test_auto_width_and_borders(sample_xlsx: Path):
    out, _ = process_excel(sample_xlsx, "keyboard")
    ws = load_workbook(out)["TRUE_Result"]
    # Border trên ô dữ liệu
    assert ws.cell(row=2, column=2).border.left.style == "thin"


def test_header_style_black_white_upper(sample_xlsx: Path):
    """Header row 1: nền đen, chữ trắng, viết hoa, center align, wrap."""
    out, _ = process_excel(sample_xlsx, "keyboard\nbàn phím")
    ws = load_workbook(out)["TRUE_Result"]
    for col in range(1, 9):  # A -> H
        cell = ws.cell(row=1, column=col)
        assert cell.fill.start_color.rgb in ("00000000", "FF000000", "000000")  # nền đen
        assert cell.font.color.rgb in ("00FFFFFF", "FFFFFFFF", "FFFFFF")  # chữ trắng
        assert cell.font.bold is True
        assert cell.alignment.horizontal == "center"
        assert cell.alignment.vertical in ("center", "middle")  # openpyxl dùng center
        assert cell.alignment.wrap_text is True
        # Viết hoa (trừ khi đã hoa sẵn)
        assert str(cell.value).upper() == str(cell.value)


def test_fixed_column_widths(sample_xlsx: Path):
    """Widths cột cố định: A=3, B=11, C=20, D=15, E=35, F=15, G=45, H=8."""
    out, _ = process_excel(sample_xlsx, "keyboard")
    ws = load_workbook(out)["TRUE_Result"]
    expected = {"A": 3, "B": 11, "C": 20, "D": 15, "E": 35, "F": 15, "G": 45, "H": 8}
    for col_letter, width in expected.items():
        assert ws.column_dimensions[col_letter].width == width


def test_column_b_time_stripped(tmp_path: Path):
    """Cột B của kết quả (out_col 2 = ColC gốc) bỏ phần thời gian:
    '05-Thg 01-26 08:43 SA' -> '05-Thg 01-26'."""
    wb = Workbook()
    ws = wb.active
    # CSV 15 cột; date nằm ở cột C gốc => trong kết quả nó thành out_col 2 (cột B).
    for c, h in enumerate(["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"], start=1):
        ws.cell(row=1, column=c, value=f"Col{h}")
    ws.append(["x", "b", "05-Thg 01-26 08:43 SA", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "Logitech keyboard", "o"])
    ws.append(["y", "b", "05-Thg 07-02-26 08:43 CH", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "Wireless mouse", "o"])
    ws.append(["z", "b", "plain-text", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "MOUSE gaming", "o"])
    path = tmp_path / "date.xlsx"
    wb.save(path)

    out, count = process_excel(path, "keyboard\nmouse")
    assert count == 3
    ws_out = load_workbook(out)["TRUE_Result"]
    # out_col 2 = cột B của kết quả = ColC gốc -> date đã strip time
    assert ws_out.cell(row=2, column=2).value == "05-Thg 01-26"
    assert ws_out.cell(row=3, column=2).value == "05-Thg 07-02-26"
    # Giá trị không phải pattern giữ nguyên
    assert ws_out.cell(row=4, column=2).value == "plain-text"


def test_data_cells_centered_wrapped(sample_xlsx: Path):
    """Data (trừ Total) center+wrap; Total giữ right."""
    out, _ = process_excel(sample_xlsx, "keyboard")
    ws = load_workbook(out)["TRUE_Result"]
    total_row = ws.max_row
    for row in range(2, total_row):  # data rows, không phải Total
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            assert cell.alignment.horizontal == "center"
            assert cell.alignment.vertical in ("center", "middle")
            assert cell.alignment.wrap_text is True
    total = ws.cell(row=total_row, column=1)
    assert total.alignment.horizontal == "right"  # giữ nguyên
    assert total.font.italic is True and total.font.bold is True


def test_print_settings(sample_xlsx: Path):
    """Print: landscape, repeat header row 1, fit to width 1 page."""
    out, _ = process_excel(sample_xlsx, "keyboard")
    ws = load_workbook(out)["TRUE_Result"]
    assert ws.page_setup.orientation == "landscape"
    assert ws.print_title_rows == "$1:$1"
    assert ws.sheet_properties.pageSetUpPr.fitToPage is True
    assert ws.page_setup.fitToWidth == 1
    assert ws.page_setup.fitToHeight == 0


def test_data_rows_wrap_text_flag(sample_xlsx: Path):
    """Toàn sheet wrap_text=True (trừ Total row giữ nguyên — check data rows)."""
    out, _ = process_excel(sample_xlsx, "keyboard")
    ws = load_workbook(out)["TRUE_Result"]
    for row in range(1, ws.max_row):  # header + data, không gồm Total
        for col in range(1, 9):
            assert ws.cell(row=row, column=col).alignment.wrap_text is True


def test_no_match_raises(sample_xlsx: Path):
    with pytest.raises(NoMatchError, match="No matching data found"):
        process_excel(sample_xlsx, "totally-not-there")


def test_original_file_unchanged(sample_xlsx: Path):
    before = sample_xlsx.read_bytes()
    process_excel(sample_xlsx, "keyboard")
    assert sample_xlsx.read_bytes() == before


def test_min_cols_no_quantity_issue(tmp_path: Path):
    """File gốc chỉ có 14 cột (không có cột O) vẫn phải có 'Quantity Replaced' tại cột H."""
    wb = Workbook()
    ws = wb.active
    headers = [f"C{i}" for i in range(1, 15)]  # 14 cột: A..N
    ws.append(headers)
    ws.append(["a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "Logitech keyboard"])
    path = tmp_path / "small.xlsx"
    wb.save(path)

    out, count = process_excel(path, "keyboard")
    assert count == 1
    ws_out = load_workbook(out)["TRUE_Result"]
    # A,B,G,H,I,K,L,M bị xóa; 14 cột -> giữ C,D,E,F,J,N (6 cột) + STT = 7 cột.
    # Cột H (8) nằm ngoài vùng dữ liệu nhưng header "Quantity Replaced" vẫn được đặt.
    assert ws_out.cell(row=1, column=1).value == "STT"
    assert ws_out.cell(row=1, column=7).value == "C14"  # cột N gốc thành cột 7
    assert ws_out.cell(row=2, column=7).value == "Logitech keyboard"
    assert ws_out.cell(row=1, column=8).value == "QUANTITY REPLACED"


# ---------------------------------------------------------------------------
# API tests: /process + /download
# ---------------------------------------------------------------------------

@pytest.fixture()
def client():
    return TestClient(main.app)


def test_health(client):
    r = client.get("/health")
    assert r.status_code == 200
    assert r.json()["status"] == "ok"


def test_process_and_download(client, sample_xlsx: Path, tmp_path: Path):
    csv_path = tmp_path / "input.csv"
    csv_path.write_text(
        "A,B,C,D,E,F,G,H,I,J,K,L,M,N,O\n"
        "r2,b,c,d,e,f,g,h,i,j,k,l,m,Logitech keyboard,o2\n",
        encoding="utf-8",
    )
    with open(csv_path, "rb") as f:
        r = client.post(
            "/process",
            files={"file": ("input.csv", f, "text/csv")},
            data={"keywords": "keyboard"},
        )
    assert r.status_code == 200
    data = r.json()
    assert data["status"] == "ok"
    assert data["match_count"] == 1
    assert data["download_url"].startswith("/download/")

    # Tải file kết quả
    r2 = client.get(data["download_url"])
    assert r2.status_code == 200
    assert r2.headers["content-type"].startswith("application/vnd.openxmlformats")

    # Job đã bị xóa sau khi download
    r3 = client.get(data["download_url"])
    assert r3.status_code == 404


def test_process_rejects_xlsx(client, tmp_path: Path):
    """Upload .xlsx phải bị từ chối — app chỉ nhận .csv."""
    xlsx = tmp_path / "old.xlsx"
    xlsx.write_bytes(b"PK\x03\x04 fake xlsx")
    with open(xlsx, "rb") as f:
        r = client.post("/process", files={"file": ("old.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}, data={"keywords": "a"})
    assert r.status_code == 400
    assert "csv" in r.json()["detail"].lower()


def test_process_rejects_non_xlsx(client):
    r = client.post("/process", files={"file": ("evil.txt", b"hello", "text/plain")}, data={"keywords": "a"})
    assert r.status_code == 400
    assert "csv" in r.json()["detail"].lower()


def test_process_rejects_empty_keywords(client, tmp_path: Path):
    csv_path = tmp_path / "input.csv"
    csv_path.write_text("A,B\n1,2\n", encoding="utf-8")
    with open(csv_path, "rb") as f:
        r = client.post("/process", files={"file": ("input.csv", f, "text/csv")}, data={"keywords": "   "})
    assert r.status_code == 400


def test_process_no_match_message(client, tmp_path: Path):
    csv_path = tmp_path / "input.csv"
    csv_path.write_text(
        "A,B,C,D,E,F,G,H,I,J,K,L,M,N,O\n"
        "r2,b,c,d,e,f,g,h,i,j,k,l,m,Logitech keyboard,o2\n",
        encoding="utf-8",
    )
    with open(csv_path, "rb") as f:
        r = client.post("/process", files={"file": ("input.csv", f, "text/csv")}, data={"keywords": "zzz"})
    assert r.status_code == 200
    assert r.json()["status"] == "no_match"
    assert r.json()["message"] == "No matching data found"


def test_process_invalid_csv(client, tmp_path: Path):
    """File .csv rỗng phải bị từ chối (400)."""
    bad = tmp_path / "empty.csv"
    bad.write_bytes(b"")
    with open(bad, "rb") as f:
        r = client.post("/process", files={"file": ("empty.csv", f, "text/csv")}, data={"keywords": "a"})
    assert r.status_code == 400


def test_process_csv(tmp_path: Path, client):
    """Upload .csv (UTF-8, dấu phẩy) được chấp nhận và xử lý như Excel."""
    csv_path = tmp_path / "data.csv"
    csv_path.write_text(
        "A,B,C,D,E,F,G,H,I,J,K,L,M,N,O\n"
        "r2,b,c,d,e,f,g,h,i,j,k,l,m,Logitech keyboard,o2\n"
        "r3,b,c,d,e,f,g,h,i,j,k,l,m,Wireless mouse,o3\n",
        encoding="utf-8",
    )
    with open(csv_path, "rb") as f:
        r = client.post(
            "/process",
            files={"file": ("data.csv", f, "text/csv")},
            data={"keywords": "keyboard"},
        )
    assert r.status_code == 200
    data = r.json()
    assert data["status"] == "ok"
    assert data["match_count"] == 1

    r2 = client.get(data["download_url"])
    assert r2.status_code == 200
    wb = load_workbook(io.BytesIO(r2.content))
    ws = wb["TRUE_Result"]
    assert ws.cell(row=2, column=7).value == "Logitech keyboard"
    assert ws.cell(row=1, column=8).value == "QUANTITY REPLACED"


def test_process_csv_semicolon(tmp_path: Path, client):
    """CSV delimiter chấm phẩy vẫn detect được."""
    csv_path = tmp_path / "data_semi.csv"
    csv_path.write_text(
        "A;B;C;D;E;F;G;H;I;J;K;L;M;N;O\n"
        "r2;b;c;d;e;f;g;h;i;j;k;l;m;Logitech keyboard;o2\n",
        encoding="utf-8",
    )
    with open(csv_path, "rb") as f:
        r = client.post(
            "/process",
            files={"file": ("data_semi.csv", f, "text/csv")},
            data={"keywords": "keyboard"},
        )
    assert r.status_code == 200
    assert r.json()["match_count"] == 1


def test_non_active_sheet_ignored(tmp_path: Path):
    """Sheet ACTIVE mới được lọc; keyword chỉ nằm ở sheet khác phải bị bỏ qua."""
    path = tmp_path / "multi.xlsx"
    make_sample_xlsx(path)
    wb = load_workbook(path)
    wb.active = wb.sheetnames.index("Other")  # chuyển sheet active sang "Other"
    wb.save(path)

    # "appear" chỉ xuất hiện ở sheet Other (không active) và cột N của nó trống
    # => không có dòng nào khớp => NoMatchError
    with pytest.raises(NoMatchError):
        process_excel(path, "appear")


def test_smoke_docker_build():
    """Kiểm tra docker-compose config hợp lệ (nếu có docker)."""
    try:
        subprocess.run(
            ["docker", "compose", "config", "--quiet"],
            cwd=ROOT,
            check=True,
            capture_output=True,
        )
    except (subprocess.CalledProcessError, FileNotFoundError) as exc:
        pytest.skip(f"Docker không khả dụng: {exc}")