"""
Xử lý Excel cho app lọc theo từ khóa.

Tái hiện macro VBA gốc:
  1. Đọc sheet đang ACTIVE của file .xlsx tải lên.
  2. Tìm các dòng mà cột N (cột 14) chứa bất kỳ từ khóa nào
     (so khớp chuỗi con, không phân biệt hoa thường).
  3. Tạo sheet mới "TRUE_Result":
       - copy dòng header từ sheet gốc
       - copy toàn bộ các dòng khớp (giá trị + format cơ bản)
       - xóa các cột A, B, G, H, I, K, L, M
       - chèn cột STT đầu tiên, đánh số 1, 2, 3...
       - đặt header "Quantity Replaced" tại cột H (cột thứ 8 của kết quả)
       - auto width toàn bộ cột + border toàn bộ bảng
  4. Không tìm thấy dòng nào -> ném NoMatchError.

Kết quả là file workbook MỚI (TRUE_Result.xlsx) — file gốc của người dùng
không bao giờ bị sửa.
"""

from copy import copy
from pathlib import Path
import io
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
# Các cột bị XÓA khỏi sheet kết quả (theo spec macro gốc).
DELETED_COLUMNS = {"A", "B", "G", "H", "I", "K", "L", "M"}

# Cột N = cột chứa từ khóa cần lọc (1-indexed).
KEYWORD_COLUMN = 14

# Cột H trong sheet kết quả (sau khi đã xóa cột và chèn STT).
QUANTITY_HEADER_COLUMN = 8

# Chiều rộng cột cố định cho sheet kết quả (theo yêu cầu in ấn).
COL_WIDTHS = {1: 3, 2: 11, 3: 20, 4: 15, 5: 35, 6: 15, 7: 45, 8: 8}

THIN = Side(style="thin", color="000000")
BOX_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Pattern cột B (out_col 2): "05-Thg 01-26 08:43 SA" -> "05-Thg 01-26".
# Bắt: prefix bất kỳ + khoảng trắng + giờ:phút + (SA|CH) tùy chọn.
DATE_TIME_RE = re.compile(r"^(.*?)\s+\d{1,2}:\d{2}(\s*(SA|CH))?\s*$", re.IGNORECASE)


class ExcelProcessingError(Exception):
    """File Excel không mở được / không hợp lệ."""


class NoMatchError(Exception):
    """Không có dòng nào chứa từ khóa."""


# ---------------------------------------------------------------------------
# CSV -> XLSX converter
# ---------------------------------------------------------------------------

def convert_csv_to_xlsx(csv_path: Path, xlsx_path: Path) -> None:
    """Chuyển file CSV thành .xlsx tạm (detect delimiter + encoding).

    Kết quả có đúng 1 sheet "Sheet1" giống như khi Excel mở CSV.
    Lỗi đọc/parse -> ExcelProcessingError.
    """
    import csv

    data = csv_path.read_bytes()
    # Detect encoding: ưu tiên UTF-8 (kể cả có BOM), fallback cp1252/latin-1.
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("cp1252", errors="replace")

    # Detect delimiter: thử dấu phẩy, chấm phẩy, tab.
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","

    try:
        rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    except Exception as exc:
        raise ExcelProcessingError(f"Không đọc được file CSV: {exc}") from exc
    if not rows:
        raise ExcelProcessingError("File CSV rỗng.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in rows:
        ws.append(row)
    wb.save(xlsx_path)


def process_excel(input_path: Path, keywords_text: str) -> tuple[Path, int]:
    """Xử lý file Excel, trả về (đường dẫn file kết quả, số dòng khớp)."""
    keywords = _parse_keywords(keywords_text)
    if not keywords:
        raise ExcelProcessingError("Vui lòng nhập ít nhất một từ khóa.")

    # Đọc 2 lần:
    #  - data_only=True  -> lấy giá trị đã tính (giống .Value của VBA) để so khớp từ khóa
    #  - data_only=False -> giữ nguyên công thức/formula khi copy sang kết quả
    try:
        wb_values = load_workbook(input_path, data_only=True)
        wb = load_workbook(input_path, data_only=False)
    except Exception as exc:
        raise ExcelProcessingError(f"Không mở được file Excel: {exc}") from exc

    src_values = wb_values.active  # sheet ACTIVE
    src = wb.active

    if src.max_row < 1:
        raise ExcelProcessingError("File Excel rỗng.")

    header_row = 1
    max_row = src.max_row
    max_col = src.max_column

    # 1) Tìm các dòng khớp: giá trị cột N chứa bất kỳ từ khóa nào.
    matched_rows: list[int] = []
    for row in range(header_row + 1, max_row + 1):
        value = src_values.cell(row=row, column=KEYWORD_COLUMN).value
        if value is None:
            continue
        text = str(value).lower()
        if any(kw in text for kw in keywords):
            matched_rows.append(row)

    if not matched_rows:
        raise NoMatchError("No matching data found")

    # 2) Các cột còn giữ lại sau khi xóa A,B,G,H,I,K,L,M (giữ thứ tự gốc).
    kept_cols = [
        col for col in range(1, max_col + 1)
        if get_column_letter(col) not in DELETED_COLUMNS
    ]

    # Layout kết quả: cột 1 = STT, cột 2.. = các cột còn giữ.
    # layout[result_col] = source_col (None cho cột STT).
    layout: list[int | None] = [None, *kept_cols]

    # 3) Tạo workbook kết quả.
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "TRUE_Result"

    # Copy dòng header.
    _copy_row(src, out_ws, header_row, layout, out_row=1)
    out_ws.cell(row=1, column=1, value="STT")

    # Copy các dòng khớp; STT dùng công thức =ROW()-1 để khi xóa hàng
    # số thứ tự tự động điều chỉnh lại (mục đích của macro gốc).
    for idx, src_row in enumerate(matched_rows, start=2):
        _copy_row(src, out_ws, src_row, layout, out_row=idx)
        out_ws.cell(row=idx, column=1, value=f"=ROW()-1")
    last_row = len(matched_rows) + 1

    # 4) Header "Quantity Replaced" tại cột H của kết quả.
    #    Nếu file gốc ít cột hơn 8, mở rộng tối thiểu tới cột H để đủ chỗ đặt header.
    out_ws.cell(row=1, column=QUANTITY_HEADER_COLUMN, value="Quantity Replaced")

    # 5) Dòng Total: merge A->G, italic + bold, align right, H = SUM(2:last_row).
    total_row = last_row + 1
    if total_row >= 2:  # luôn đúng vì có >=1 dòng khớp (đã check NoMatchError)
        out_ws.merge_cells(
            start_row=total_row, start_column=1, end_row=total_row, end_column=7
        )
        total_cell = out_ws.cell(row=total_row, column=1, value="Total")
        total_cell.font = Font(italic=True, bold=True)
        total_cell.alignment = Alignment(horizontal="right")
        sum_cell = out_ws.cell(
            row=total_row,
            column=QUANTITY_HEADER_COLUMN,
            value=f"=SUM(H2:H{last_row})",
        )
        sum_cell.font = Font(italic=True, bold=True)

    # 6) Copy vùng merged (nếu header bị merge) để giữ giao diện.
    _copy_merges(src, out_ws, matched_rows, layout, header_row)

    # 7) Format nâng cao: header (đen/trắng/hoa/giữa), width cột cố định,
    #    strip thời gian cột B, alignment toàn sheet (trừ Total), wrap text,
    #    border, print settings.
    _apply_header_style(out_ws)
    _apply_fixed_widths(out_ws)
    _strip_column_b_times(out_ws, last_row)
    _apply_alignment(out_ws, total_row)
    _apply_borders(out_ws, total_row)
    _apply_print_settings(out_ws)

    # 8) Lưu file kết quả cạnh file gốc (cùng job dir -> dễ dọn dẹp).
    output_path = input_path.with_name("TRUE_Result.xlsx")
    out_wb.save(output_path)

    return output_path, len(matched_rows)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_keywords(text: str) -> list[str]:
    """Tách từ khóa: mỗi dòng 1 từ khóa, bỏ dòng trống, lowercase để so khớp."""
    return [line.strip().lower() for line in text.splitlines() if line.strip()]


def _copy_row(src_ws, out_ws, src_row: int, layout, out_row: int) -> None:
    """Copy 1 dòng từ sheet gốc sang sheet kết quả theo layout cột."""
    for out_col, src_col in enumerate(layout, start=1):
        if src_col is None:
            continue
        _copy_cell(src_ws.cell(row=src_row, column=src_col),
                   out_ws.cell(row=out_row, column=out_col))


def _copy_cell(src_cell, dst_cell) -> None:
    """Copy giá trị + format cơ bản (font, fill, border, alignment, number format).

    Lưu ý: một số cell đọc từ file có thể có style thuộc tính là None
    (openpyxl không luôn luôn trả PatternFill/Border) — phải guard trước khi copy.
    """
    if src_cell.value is not None:
        dst_cell.value = src_cell.value
    dst_cell.number_format = src_cell.number_format
    if src_cell.has_style:
        for attr in ("font", "fill", "border", "alignment", "protection"):
            style = getattr(src_cell, attr, None)
            if style is not None:
                try:
                    setattr(dst_cell, attr, copy(style))
                except TypeError:
                    # Style không copy được (hiếm) — bỏ qua, không chặn xử lý.
                    continue


def _copy_merges(src_ws, out_ws, matched_rows, layout, header_row) -> None:
    """Copy vùng merged nằm trong các dòng đã copy (chủ yếu cho header)."""
    out_row_of = {header_row: 1}
    for idx, src_r in enumerate(matched_rows, start=2):
        out_row_of[src_r] = idx

    kept_set = {sc for sc in layout if sc is not None}
    out_col_of = {sc: oc for oc, sc in enumerate(layout) if sc}

    for rng in src_ws.merged_cells.ranges:
        rows_in = sorted(r for r in range(rng.min_row, rng.max_row + 1) if r in out_row_of)
        cols_in = [c for c in range(rng.min_col, rng.max_col + 1) if c in kept_set]
        if not rows_in or not cols_in:
            continue
        row_min = out_row_of[rows_in[0]]
        row_max = out_row_of[rows_in[-1]]
        col_min = out_col_of[cols_in[0]]
        col_max = out_col_of[cols_in[-1]]
        if row_min == row_max and col_min == col_max:
            continue
        out_ws.merge_cells(start_row=row_min, start_column=col_min,
                           end_row=row_max, end_column=col_max)


def _apply_borders(ws, last_row: int) -> None:
    """Border mảnh màu đen cho toàn bộ bảng (header + dữ liệu)."""
    if last_row < 1 or ws.max_column < 1:
        return
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = BOX_BORDER


def _auto_width(ws, last_row: int) -> None:
    """(Không còn dùng) — width cột đã cố định theo COL_WIDTHS."""
    pass


def _apply_header_style(ws) -> None:
    """Header row 1: nền đen, chữ trắng, đậm, viết hoa, center align, wrap."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="000000")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, 9):  # A -> H
        cell = ws.cell(row=1, column=col)
        if cell.value is not None:
            cell.value = str(cell.value).upper()
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align


def _apply_fixed_widths(ws) -> None:
    """Set width cố định cho 8 cột kết quả theo COL_WIDTHS."""
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _strip_time_from_date(value):
    """Bỏ phần giờ:phút + (SA|CH) khỏi chuỗi nếu khớp pattern date+time."""
    if not isinstance(value, str):
        return value
    m = DATE_TIME_RE.match(value.strip())
    if m:
        return m.group(1).strip()
    return value


def _strip_column_b_times(ws, last_row: int) -> None:
    """Cột B (out_col=2) của kết quả: bỏ phần thời gian ở mỗi ô dữ liệu."""
    for row in range(2, last_row + 1):
        cell = ws.cell(row=row, column=2)
        cell.value = _strip_time_from_date(cell.value)


def _apply_alignment(ws, total_row: int) -> None:
    """Toàn sheet (trừ dòng Total) center align + wrap text;
    dòng Total giữ nguyên align right (không đổi style của nó)."""
    for row in range(1, total_row):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_print_settings(ws) -> None:
    """Print setup: landscape, A4, repeat header row 1, fit 1 trang ngang."""
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_title_rows = "1:1"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0